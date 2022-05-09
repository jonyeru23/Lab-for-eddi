import json
import openpyxl
"""
This is a script to take the parameters from Eddington and put them in the Excel file of your choice!
You might need to install the libraries above for it to work' after that just click on the file and let it run
"""

def main():
    json_path = input('Enter Json path:')[1:-1]
    exel_path = input('Enter excel path:')[1:-1]

    try:
        write_to_excel(make_dict(jason_path=json_path), exel_path)
    except:
        raise Exception('Expected path of shape: "C:\lablabla\your_file"')

def make_dict(jason_path):
    f = open(jason_path)
    return json.load(f)

def write_to_excel(data, file_path):
    headers = ['a', 'delta_a', 'rel_error_a']
    keys = ['a', "aerr", "arerr"]

    wb = openpyxl.load_workbook(filename=file_path)
    wb.create_sheet('Eddi params')

    ws = wb['Eddi params']

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for col, key in enumerate(keys, start=1):
        for row, value in enumerate(data[key], start=2):
            ws.cell(row=row, column=col, value=value)

    wb.save(file_path)

if __name__ == '__main__':
    main()

